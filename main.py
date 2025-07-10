from flask import Flask, request, jsonify
import os
from openpyxl import Workbook, load_workbook

UPLOAD_FOLDER = 'selfies'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def ensure_excel_headers(file_path):
    headers_needed = ['Имя', 'Почта', 'Пароль', 'Путь к фото']
    wb = load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    changed = False
    # Добавить недостающие заголовки
    for h in headers_needed:
        if h not in headers:
            ws.cell(row=1, column=len(headers)+1, value=h)
            headers.append(h)
            changed = True
    # Убедиться, что у всех строк есть нужное количество ячеек
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if len(row) < len(headers_needed):
            for j in range(len(row)+1, len(headers_needed)+1):
                ws.cell(row=i, column=j, value='')
    if changed:
        wb.save(file_path)

app = Flask(__name__)

@app.route('/')
def index():
    return 'Telegram Mini App backend is running!'

@app.route('/api/register', methods=['POST'])
def register():
    data = request.json
    name = data.get('name')
    email = data.get('email')
    password = data.get('password')
    if not name or not email or not password:
        return jsonify({'status': 'error', 'error': 'Все поля обязательны!'}), 400

    file_path = 'users.xlsx'
    if os.path.exists(file_path):
        ensure_excel_headers(file_path)
    else:
        # Создаём файл и добавляем заголовки
        wb = Workbook()
        ws = wb.active
        ws.append(['Имя', 'Почта', 'Пароль', 'Путь к фото'])
        wb.save(file_path)

    # Добавляем пользователя
    wb = load_workbook(file_path)
    ws = wb.active
    ws.append([name, email, password, ''])
    wb.save(file_path)
    return jsonify({'status': 'ok'})

@app.route('/api/upload_selfie', methods=['POST'])
def upload_selfie():
    email = request.form.get('email')
    photo = request.files.get('photo')
    if not email or not photo:
        return jsonify({'status': 'error', 'error': 'Нет email или фото'}), 400

    # Сохраняем фото
    filename = f'{email.replace("@", "_").replace(".", "_")}.jpg'
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    photo.save(filepath)

    # Сохраняем путь к фото в Excel
    file_path = 'users.xlsx'
    if os.path.exists(file_path):
        ensure_excel_headers(file_path)
    wb = load_workbook(file_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    photo_col = headers.index('Путь к фото') + 1
    for row in ws.iter_rows(min_row=2):
        if str(row[1].value).strip().lower() == email.lower():
            row[photo_col-1].value = filepath
            break
    else:
        return jsonify({'status': 'error', 'error': 'Пользователь не найден'}), 404
    wb.save(file_path)
    return jsonify({'status': 'ok'})

@app.route('/api/procedure', methods=['POST'])
def add_procedure():
    data = request.json
    print('Получена процедура:', data)
    return jsonify({'status': 'ok'})

if __name__ == '__main__':
    app.run(debug=True)
